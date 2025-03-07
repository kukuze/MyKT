import http

import cloudscraper
import openai
import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook

from constant import Constants


def getContentAndTags(cid, qindex):
    url = f"https://codeforces.com/contest/{cid}/problem/{qindex}"
    print(url)
    scraper = cloudscraper.create_scraper()
    included_text = None
    diff = None
    tags = []
    try:
        response = scraper.get(url, timeout=10)
        response.raise_for_status()  # Raise an exception for bad responses
        html_content = response.text
        soup = BeautifulSoup(html_content, 'html.parser')

        # Exclude specific tags with certain attributes
        excluded_tag_attrs = {"class": "sample-tests"}
        excluded_tag_name = "div"
        excluded_tag_attrs2 = {"class": "note"}
        excluded_tag_name2 = "div"

        # Find and decompose excluded tags
        excluded_tag = soup.find(excluded_tag_name, excluded_tag_attrs)
        if excluded_tag:
            excluded_tag.decompose()

        excluded_tag = soup.find(excluded_tag_name2, excluded_tag_attrs2)
        if excluded_tag:
            excluded_tag.decompose()

        # Find included tag for problem statement
        included_tag_attrs = {"class": "problem-statement"}
        included_tag_name = "div"
        included_tag = soup.find(included_tag_name, included_tag_attrs)

        # Get text from included tag
        if included_tag:
            included_text = included_tag.get_text(separator=' ', strip=True)
        else:
            included_text = None

        # Find and extract tags

        tag_elements = soup.find_all("span", class_="tag-box")
        for tag_element in tag_elements:
            text = tag_element.get_text(strip=True)
            if text == "*special problem":  ## 在api中是*special 页面中是 *special problem 统一处理
                tags.append("*special")
                continue
            if text.startswith("*"):  # 忽略掉 *800这种分数标签
                diff = text[1:]
            tags.append(text)

        tags_str = ",".join(tags) if tags else None
        return included_text, tags_str, diff

    except requests.exceptions.RequestException as e:
        print(f"Error fetching content from {url}: {e}")
        return None, None, None
    finally:
        scraper.close()


# def from_excel_content_html():
#     global index, cid, qindex, type
#     # 读取 Excel 文件
#     file_path = r"D:\MyKT\data\new.xlsx"
#     df = pd.read_excel(file_path)
#     # 加载工作簿
#     workbook = load_workbook(file_path)
#     sheet = workbook.active  # 获取活动工作表
#     # 筛选出 content 或 difficulty 为空或为 0，或者 tags 为空的行
#     filtered_df = df[(df['content'].isnull())]
#     # 遍历筛选后的行，调用函数并更新 Excel 文件逐行保存
#     for index, row in filtered_df.iterrows():
#         cid = row['cid']  # 假设 Excel 中有 'cid' 列
#         qindex = row['qindex']  # 假设 Excel 中有 'qindex' 列
#         type = row['type']
#         if cid > 1800:
#             included_text, _, _ = getContentAndTags(cid, qindex)  # 调用函数获取返回值
#
#             # 更新 DataFrame 中的对应行（仅更新成功获取到的值）
#             if included_text is not None:  # 如果获取到了 content，则更新
#                 df.at[index, 'content'] = included_text
#                 sheet.cell(row=index + 2, column=df.columns.get_loc('content') + 1, value=included_text)  # 更新 Excel 文件
#                 workbook.save(file_path)
#                 print(included_text)
#         # if tags_str is not None:  # 如果获取到了 tags，则更新
#         #     df.at[index, 'tags'] = tags_str
#         #     sheet.cell(row=index + 2, column=df.columns.get_loc('tags') + 1, value=tags_str)  # 更新 Excel 文件
#         #
#         # if diff is not None:  # 如果获取到了 difficulty，则更新
#         #     df.at[index, 'difficulty'] = diff
#         #     sheet.cell(row=index + 2, column=df.columns.get_loc('difficulty') + 1, value=diff)  # 更新 Excel 文件
#
#         # 每更新一行，保存到文件
#     # 关闭工作簿
#     workbook.close()


import os
import requests
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
import json
import cloudscraper

def download_pdf(url, save_path):
    """
    下载 PDF 文件并保存到本地
    :param url: PDF 文件的 URL
    :param save_path: 保存 PDF 文件的路径
    """
    scraper = cloudscraper.create_scraper()
    try:
        response = scraper.get(url, stream=True)
        response.raise_for_status()  # 检查请求是否成功
        with open(save_path, "wb") as pdf_file:
            for chunk in response.iter_content(chunk_size=1024):
                pdf_file.write(chunk)
        print(f"PDF downloaded successfully and saved to {save_path}")
    except Exception as e:
        print(f"Error downloading PDF: {e}")

def read_pdf(file_path):
    """
    读取 PDF 文件中的文字内容
    :param file_path: 本地 PDF 文件路径
    :return: PDF 文件中的文字内容
    """
    try:
        with pdfplumber.open(file_path) as pdf:
            text = ""
            for page in pdf.pages:
                text += page.extract_text() + "\n"  # 提取每一页的文字
            return text
    except Exception as e:
        print(f"Error reading PDF: {e}")
        return None

def extract_text_with_openai(prompt):
    """
    使用 OpenAI 接口提取文本
    :param prompt: 提示词
    :return: 提取的文本
    """
    messages = [
        {"role": "system", "content": "You are a helpful assistant. You can help people organize the text of programming competition questions."}
    ]
    try:
        client = openai.OpenAI(
            base_url="https://yunwu.ai/v1",
            api_key=Constants.openai_api_key
        )
        messages.append({"role": "user", "content": prompt})
        response = client.chat.completions.create(
            model="chatgpt-4o-latest",
            messages=messages,
            temperature=1
        )

        assistant_response = response.choices[0].message.content
        return assistant_response
    except Exception as e:
        print(f"Error calling OpenAI API: {e}")
        return None


def update_excel(file_path, df, index, column, value):
    """
    更新 Excel 文件中的指定单元格
    :param file_path: Excel 文件路径
    :param df: DataFrame 对象
    :param index: 行索引
    :param column: 列名
    :param value: 更新的值
    """
    try:
        # 更新 DataFrame
        df.at[index, column] = value

        # 更新 Excel 文件
        workbook = load_workbook(file_path)
        sheet = workbook.active
        sheet.cell(row=index + 2, column=df.columns.get_loc(column) + 1, value=value)
        workbook.save(file_path)
        workbook.close()
        print(f"Updated Excel at row {index + 2}, column {column} with value: {value}")
    except Exception as e:
        print(f"Error updating Excel: {e}")

# 主程序
def main():
    # Excel 文件路径
    file_path = r"D:\MyKT\data\codeforces_problem_detail.xlsx"

    # 读取 Excel 文件
    df = pd.read_excel(file_path)

    # 筛选出 content 为空的行
    filtered_df = df[df['content'].isnull()]

    # 遍历筛选后的行
    for index, row in filtered_df.iterrows():
        cid = row['cid']  # 假设 Excel 中有 'cid' 列
        qindex = row['qindex']  # 假设 Excel 中有 'qindex' 列
        pdf_url = f"https://codeforces.com/contest/{cid}/problem/{qindex}"  # 替换为实际的 PDF URL 模板
        save_path = f"problem_{cid}_{qindex}.pdf"  # 保存的文件名
        print(pdf_url)
        # 判断文件是否已经存在
        if not os.path.exists(save_path):
            print(f"PDF not found locally. Downloading: {save_path}")
            # 下载 PDF
            download_pdf(pdf_url, save_path)
        else:
            print(f"PDF already exists: {save_path}")

        # 读取 PDF 内容
        pdf_text = read_pdf(save_path)
        if pdf_text:
            # 使用 OpenAI 提取题目文本
            prompt = f"Only extract the problem description within quotation marks. If there is any unreasonable format, please organize it appropriately. Ignore the input and output parts and do not use markdown format.'{pdf_text}'"
            extracted_text = extract_text_with_openai(prompt)

            if extracted_text:
                # 更新到 Excel
                update_excel(file_path, df, index, 'content', extracted_text)

        # # 删除临时 PDF 文件
        # if os.path.exists(save_path):
        #     os.remove(save_path)
def from_json_to_excel():
    # 读取 JSON 文件
    json_file_path = r"D:\MyKT\data\no_content_set.json"
    with open(json_file_path, 'r', encoding='utf-8') as f:
        no_content_data = json.load(f)
    print(no_content_data)
    # # 创建一个新的 Excel 文件
    # output_file_path = r"D:\MyKT\data\content_filled.xlsx"
    # workbook = Workbook()
    # sheet = workbook.active
    # sheet.title = "Content Data"
    #
    # # 写入表头
    # sheet.append(["cid", "qindex", "content"])
    #
    # # 遍历 JSON 数据，获取 content 并写入 Excel
    # for item in no_content_data:
    #     cid = item[0]  # 获取 cid
    #     qindex = item[1]  # 获取 qindex
    #
    #     # 调用 getContentAndTags 函数获取 content
    #     content, _, _ = getContentAndTags(cid, qindex)
    #     if content:
    #         # 写入到 Excel
    #         sheet.append([cid, qindex, content])
    #
    # # 保存 Excel 文件
    # workbook.save(output_file_path)
    # print(f"Content-filled Excel file saved to: {output_file_path}")


import json
import pandas as pd


def from_json_to_excel_supplement():
    # 读取 JSON 文件
    json_file_path = r"D:\MyKT\data\no_content_set.json"
    with open(json_file_path, 'r', encoding='utf-8') as f:
        no_content_data = json.load(f)

    # 读取现有的 Excel 文件
    excel_file_path = r"D:\MyKT\data\content_filled.xlsx"
    try:
        workbook = load_workbook(excel_file_path)
        sheet = workbook.active
    except FileNotFoundError:
        print(f"Excel 文件未找到: {excel_file_path}")
        return

    # 获取现有 Excel 数据到 DataFrame
    existing_data = pd.DataFrame(sheet.values)
    existing_data.columns = existing_data.iloc[0]  # 第一行作为列名
    existing_data = existing_data[1:]  # 去掉表头行
    existing_data = existing_data.reset_index(drop=True)

    # 遍历 JSON 数据，检查并添加缺失内容
    for item in no_content_data:
        cid = item[0]  # 获取 cid
        qindex = item[1]  # 获取 qindex

        # 检查是否已存在于 Excel 中
        if not ((existing_data["cid"] == cid) & (existing_data["qindex"] == qindex)).any():
            # 调用 getContentAndTags 函数获取 content
            content, _, _ = getContentAndTags(cid, qindex)
            if content:
                # 添加到 Excel
                sheet.append([cid, qindex, content])
                print(f"添加到 Excel: cid={cid}, qindex={qindex}, content={content}")

    # 保存 Excel 文件
    workbook.save(excel_file_path)
    print(f"更新后的 Excel 文件已保存到: {excel_file_path}")


import pandas as pd


import pandas as pd

def merge_excel_content():
    # 读取两个Excel文件
    content_df = pd.read_excel('D:/MyKT/data/content_filled.xlsx')  # 只有 cid, qindex, content 三列
    problem_df = pd.read_excel('D:/MyKT/data/codeforces_problem_detail.xlsx')  # 包含更多列

    # 确保只保留 problem_df 中的 cid, qindex, content 列
    problem_df = problem_df[['cid', 'qindex', 'content']]

    # 合并两个 DataFrame，基于 cid 和 qindex 列
    merged_df = pd.merge(content_df, problem_df, on=['cid', 'qindex'], how='outer', suffixes=('_content', '_problem'))

    # 如果需要，可以对合并后的 content 列进行处理，例如优先使用 content_filled 的内容
    merged_df['content'] = merged_df['content_content'].combine_first(merged_df['content_problem'])

    # 删除多余的列（如果不需要保留原始的 content_content 和 content_problem）
    merged_df.drop(columns=['content_content', 'content_problem'], inplace=True)

    # 保存合并后的结果到新的 Excel 文件
    merged_df.to_excel('D:/MyKT/data/merged_content.xlsx', index=False)

    print("合并完成，结果已保存到 'D:/MyKT/data/merged_content.xlsx'")




if __name__ == "__main__":
    merge_excel_content()


